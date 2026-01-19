#!/usr/bin/env python3
"""
Home Health Analytics API Server

FastAPI server that provides analytics data to the Next.js frontend.
Serves all pivot tables and summary metrics as JSON endpoints.
"""

from fastapi import FastAPI, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
import pandas as pd
import json
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional
import uvicorn
from pydantic import BaseModel

# Import our analytics modules
from home_health_extractor import HomeHealthExtractor
from pivot_analytics import HomeHealthAnalytics
from profitability_analysis import ProfitabilityAnalyzer
from fastapi.responses import FileResponse

# Import authentication
from auth import (
    login as auth_login,
    get_current_user,
    User,
    LoginRequest,
    Token
)

# Check if running in production
IS_PRODUCTION = os.getenv("ENVIRONMENT", "development").lower() == "production"

# Disable OpenAPI docs in production for security
app = FastAPI(
    title="Home Health Analytics API",
    description="API for home health billing analytics and pivot tables",
    version="1.0.0",
    docs_url=None if IS_PRODUCTION else "/docs",
    redoc_url=None if IS_PRODUCTION else "/redoc",
    openapi_url=None if IS_PRODUCTION else "/openapi.json"
)

# CORS Configuration
# In production, set CORS_ORIGINS environment variable with comma-separated allowed origins
# Example: CORS_ORIGINS=https://yourdomain.com,https://app.yourdomain.com
cors_origins_env = os.getenv("CORS_ORIGINS", "")
if cors_origins_env:
    CORS_ORIGINS = [origin.strip() for origin in cors_origins_env.split(",")]
else:
    # Development defaults
    CORS_ORIGINS = ["http://localhost:3000", "http://localhost:3001", "http://localhost:3100"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Global variables to cache data
analytics_data = {}
summary_metrics = {}
last_updated = None

class ProcessPDFRequest(BaseModel):
    pdf_directory: Optional[str] = "data/pdfs"

def load_latest_analytics():
    """Load the latest analytics data."""
    global analytics_data, summary_metrics, last_updated

    try:
        # Find the most recent analytics file
        analytics_files = list(Path("analytics_output").glob("home_health_analytics_*.xlsx"))
        if not analytics_files:
            return False

        latest_file = max(analytics_files, key=lambda x: x.stat().st_mtime)
        last_updated = datetime.fromtimestamp(latest_file.stat().st_mtime)

        # Load all sheets
        analytics_data = {}
        excel_file = pd.ExcelFile(latest_file)

        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(latest_file, sheet_name=sheet_name)
            # Convert datetime columns to strings for JSON serialization
            for col in df.columns:
                if df[col].dtype == 'datetime64[ns]':
                    df[col] = df[col].dt.strftime('%Y-%m-%d').fillna('')

            # Replace NaN values with 0 for numeric columns or empty string for others
            df = df.fillna(0)

            # Convert any remaining NaN to 0 (for numeric) or empty string
            df = df.replace([float('inf'), float('-inf')], 0)

            analytics_data[sheet_name] = df.to_dict('records')

        # Calculate summary metrics
        if 'Original Claims' in analytics_data:
            claims_df = pd.DataFrame(analytics_data['Original Claims'])
            visits_df = pd.DataFrame(analytics_data['Original Visits'])
            master_df = pd.DataFrame(analytics_data['Master Transformed'])

            summary_metrics = {
                'total_patients': claims_df['Patient Name'].nunique() if not claims_df.empty else 0,
                'total_claims': len(claims_df),
                'total_visits': len(visits_df),
                'total_billed': float(claims_df['Total Amount'].sum()) if not claims_df.empty else 0,
                'total_collected': float(claims_df['Posted Payments'].sum()) if not claims_df.empty else 0,
                'total_outstanding': float(claims_df['Balance'].sum()) if not claims_df.empty else 0,
                'collection_rate': float((claims_df['Posted Payments'].sum() / claims_df['Total Amount'].sum() * 100)) if not claims_df.empty and claims_df['Total Amount'].sum() > 0 else 0,
                'avg_claim_amount': float(claims_df['Total Amount'].mean()) if not claims_df.empty else 0,
                'total_service_cost': float(master_df['Total Cost'].sum()) if not master_df.empty else 0,
                'last_updated': last_updated.isoformat()
            }

            summary_metrics['gross_profit'] = summary_metrics['total_collected'] - summary_metrics['total_service_cost']
            summary_metrics['profit_margin'] = (summary_metrics['gross_profit'] / summary_metrics['total_collected'] * 100) if summary_metrics['total_collected'] > 0 else 0

        return True

    except Exception as e:
        print(f"Error loading analytics: {e}")
        return False

@app.on_event("startup")
async def startup_event():
    """Load analytics data on startup."""
    load_latest_analytics()

@app.get("/")
async def root():
    """Root endpoint with API information."""
    # In production, don't expose endpoint list
    if IS_PRODUCTION:
        return {
            "message": "Home Health Analytics API",
            "version": "1.0.0",
            "status": "running"
        }

    return {
        "message": "Home Health Analytics API",
        "version": "1.0.0",
        "endpoints": [
            "/analytics/summary",
            "/analytics/revenue-by-claim",
            "/analytics/service-costs",
            "/analytics/profitability-by-patient",
            "/analytics/provider-performance",
            "/analytics/code-performance",
            "/analytics/service-cost-summary",
            "/analytics/insurance-performance"
        ],
        "last_updated": last_updated.isoformat() if last_updated else None
    }


# ============ AUTHENTICATION ENDPOINTS ============

@app.post("/auth/login", response_model=Token)
async def login(request: LoginRequest):
    """
    Authenticate user and return JWT token.

    Default credentials (for development only):
    - Username: admin
    - Password: homehealth2024

    In production, set AUTH_USERNAME and AUTH_PASSWORD environment variables.
    """
    token = auth_login(request.username, request.password)

    if not token:
        raise HTTPException(
            status_code=401,
            detail="Invalid username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )

    return token


@app.get("/auth/verify")
async def verify_auth(current_user: User = Depends(get_current_user)):
    """Verify that the current token is valid."""
    return {"authenticated": True, "username": current_user.username}


@app.get("/auth/me")
async def get_me(current_user: User = Depends(get_current_user)):
    """Get current user information."""
    return {"username": current_user.username}

@app.get("/analytics/summary")
async def get_summary(current_user: User = Depends(get_current_user)):
    """Get summary metrics for the dashboard."""
    if not summary_metrics:
        load_latest_analytics()

    if not summary_metrics:
        raise HTTPException(status_code=404, detail="No analytics data available")

    return summary_metrics

@app.get("/analytics/revenue-by-claim")
async def get_revenue_by_claim(current_user: User = Depends(get_current_user)):
    """Get revenue by claim analysis."""
    if 'Revenue by Claim' not in analytics_data:
        load_latest_analytics()

    if 'Revenue by Claim' not in analytics_data:
        raise HTTPException(status_code=404, detail="Revenue by claim data not available")

    return analytics_data['Revenue by Claim']

@app.get("/analytics/service-costs")
async def get_service_costs(current_user: User = Depends(get_current_user)):
    """Get service costs analysis."""
    if 'Service Costs' not in analytics_data:
        load_latest_analytics()

    if 'Service Costs' not in analytics_data:
        raise HTTPException(status_code=404, detail="Service costs data not available")

    return analytics_data['Service Costs']

@app.get("/analytics/profitability-by-patient")
async def get_profitability_by_patient(current_user: User = Depends(get_current_user)):
    """Get profitability by patient analysis."""
    if 'Profitability by Patient' not in analytics_data:
        load_latest_analytics()

    if 'Profitability by Patient' not in analytics_data:
        raise HTTPException(status_code=404, detail="Profitability data not available")

    return analytics_data['Profitability by Patient']

@app.get("/analytics/provider-performance")
async def get_provider_performance(current_user: User = Depends(get_current_user)):
    """Get provider performance analysis."""
    if 'Provider Performance' not in analytics_data:
        load_latest_analytics()

    if 'Provider Performance' not in analytics_data:
        raise HTTPException(status_code=404, detail="Provider performance data not available")

    return analytics_data['Provider Performance']

@app.get("/analytics/code-performance")
async def get_code_performance(current_user: User = Depends(get_current_user)):
    """Get claim code performance analysis."""
    if 'Code Performance' not in analytics_data:
        load_latest_analytics()

    if 'Code Performance' not in analytics_data:
        raise HTTPException(status_code=404, detail="Code performance data not available")

    return analytics_data['Code Performance']

@app.get("/analytics/service-cost-summary")
async def get_service_cost_summary(current_user: User = Depends(get_current_user)):
    """Get service cost summary."""
    if 'Service Cost Summary' not in analytics_data:
        load_latest_analytics()

    if 'Service Cost Summary' not in analytics_data:
        raise HTTPException(status_code=404, detail="Service cost summary not available")

    return analytics_data['Service Cost Summary']

@app.get("/analytics/insurance-performance")
async def get_insurance_performance(current_user: User = Depends(get_current_user)):
    """Get insurance payer performance."""
    if 'Insurance Payer Performance' not in analytics_data:
        load_latest_analytics()

    if 'Insurance Payer Performance' not in analytics_data:
        raise HTTPException(status_code=404, detail="Insurance performance data not available")

    return analytics_data['Insurance Payer Performance']

@app.get("/analytics/patient/{patient_name}")
async def get_patient_details(patient_name: str, current_user: User = Depends(get_current_user)):
    """Get detailed information for a specific patient."""
    if not analytics_data:
        load_latest_analytics()

    patient_data = {}

    try:
        # Get patient claims
        if 'Original Claims' in analytics_data:
            claims = [claim for claim in analytics_data['Original Claims']
                     if claim['Patient Name'].lower() == patient_name.lower()]
            patient_data['claims'] = claims

        # Get patient visits
        if 'Original Visits' in analytics_data:
            visits = [visit for visit in analytics_data['Original Visits']
                     if visit.get('Patient Name Clean', '').lower() == patient_name.lower()]
            patient_data['visits'] = visits

        # Get patient service records
        if 'Master Transformed' in analytics_data:
            services = [service for service in analytics_data['Master Transformed']
                       if service['Patient Name'].lower() == patient_name.lower()]
            patient_data['services'] = services

        # Get patient profitability
        if 'Profitability by Patient' in analytics_data:
            profitability = [profit for profit in analytics_data['Profitability by Patient']
                           if profit['Patient Name'].lower() == patient_name.lower()]
            patient_data['profitability'] = profitability

        if not any(patient_data.values()):
            raise HTTPException(status_code=404, detail=f"Patient '{patient_name}' not found")

        return patient_data

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving patient data: {str(e)}")

@app.post("/process-pdfs")
async def process_pdfs(request: ProcessPDFRequest, current_user: User = Depends(get_current_user)):
    """Process PDFs and regenerate analytics."""
    try:
        # Extract data from PDFs
        extractor = HomeHealthExtractor()
        claims_df, visits_df = extractor.process_all_pdfs(request.pdf_directory)

        if claims_df.empty and visits_df.empty:
            raise HTTPException(status_code=400, detail="No data extracted from PDFs")

        # Save extracted data
        extractor.save_extracted_data(claims_df, visits_df)

        # Generate analytics
        analytics_engine = HomeHealthAnalytics()
        analytics = analytics_engine.generate_all_analytics(claims_df, visits_df)
        analytics_engine.save_analytics(analytics)

        # Reload analytics data
        load_latest_analytics()

        return {
            "message": "PDFs processed successfully",
            "claims_extracted": len(claims_df),
            "visits_extracted": len(visits_df),
            "analytics_generated": True,
            "last_updated": datetime.now().isoformat()
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing PDFs: {str(e)}")

@app.get("/analytics/refresh")
async def refresh_analytics(current_user: User = Depends(get_current_user)):
    """Refresh analytics data from the latest files."""
    success = load_latest_analytics()

    if not success:
        raise HTTPException(status_code=404, detail="No analytics data found")

    return {
        "message": "Analytics data refreshed",
        "last_updated": last_updated.isoformat() if last_updated else None,
        "tables_loaded": list(analytics_data.keys())
    }


# ============ PROFITABILITY ENDPOINTS ============

@app.get("/profitability/analysis")
async def get_profitability_analysis(current_user: User = Depends(get_current_user)):
    """Get complete profitability analysis from Claims and Employee Costs."""
    try:
        analyzer = ProfitabilityAnalyzer()
        results = analyzer.analyze()

        if "error" in results:
            raise HTTPException(status_code=500, detail=results["error"])

        return results
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error analyzing profitability: {str(e)}")


@app.get("/profitability/overall")
async def get_profitability_overall(current_user: User = Depends(get_current_user)):
    """Get overall profitability summary."""
    try:
        analyzer = ProfitabilityAnalyzer()
        results = analyzer.analyze()
        return results.get("overall", {})
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/profitability/by-physician")
async def get_profitability_by_physician(current_user: User = Depends(get_current_user)):
    """Get profitability breakdown by physician."""
    try:
        analyzer = ProfitabilityAnalyzer()
        results = analyzer.analyze()
        return results.get("by_physician", [])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/profitability/unmatched")
async def get_unmatched_patients(current_user: User = Depends(get_current_user)):
    """Get unmatched patient costs (data inconsistencies)."""
    try:
        analyzer = ProfitabilityAnalyzer()
        results = analyzer.analyze()
        return results.get("unmatched_patients", [])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/profitability/overhead")
async def get_overhead_costs(current_user: User = Depends(get_current_user)):
    """Get overhead/admin costs."""
    try:
        analyzer = ProfitabilityAnalyzer()
        results = analyzer.analyze()
        return results.get("overhead", [])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/profitability/export")
async def export_profitability_excel(current_user: User = Depends(get_current_user)):
    """Export profitability analysis to Excel and return the file."""
    try:
        analyzer = ProfitabilityAnalyzer()
        output_path = analyzer.export_to_excel()

        return FileResponse(
            path=output_path,
            filename=Path(output_path).name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting: {str(e)}")


# ============ DATA EXPLORER ENDPOINTS ============

@app.get("/explorer/claims")
async def get_claims_data(current_user: User = Depends(get_current_user)):
    """Get all claims data for exploration."""
    try:
        claims = pd.read_csv("data/excel/Claim List.csv")

        # Parse dates
        date_cols = ['SOC Date', 'Claim Start', 'Claim End', 'RAP Sent Date', 'Final Sent Date']
        for col in date_cols:
            if col in claims.columns:
                claims[col] = pd.to_datetime(claims[col], errors='coerce').dt.strftime('%Y-%m-%d')

        # Fill NaN
        claims = claims.fillna('')

        return {
            "data": claims.to_dict('records'),
            "columns": claims.columns.tolist(),
            "total_records": len(claims)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/explorer/costs")
async def get_costs_data(current_user: User = Depends(get_current_user)):
    """Get all employee costs data for exploration."""
    try:
        costs = pd.read_excel("data/excel/employee_costs.xlsx")

        # Parse dates
        costs['Date'] = pd.to_datetime(costs['Date'], errors='coerce').dt.strftime('%Y-%m-%d')
        costs['Date_Paid'] = pd.to_datetime(costs['Date_Paid'], errors='coerce').dt.strftime('%Y-%m-%d')

        # Fill NaN
        costs = costs.fillna('')

        return {
            "data": costs.to_dict('records'),
            "columns": costs.columns.tolist(),
            "total_records": len(costs)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/explorer/monthly-summary")
async def get_monthly_summary(current_user: User = Depends(get_current_user)):
    """Get monthly revenue and costs summary for charts."""
    try:
        claims = pd.read_csv("data/excel/Claim List.csv")
        costs = pd.read_excel("data/excel/employee_costs.xlsx")

        # Parse claim dates and group by month
        claims['Claim_Month'] = pd.to_datetime(claims['Claim Start'], errors='coerce').dt.to_period('M')
        claims_monthly = claims.groupby('Claim_Month').agg({
            'Claim Amount': 'sum',
            'Paid Amount': 'sum',
            'Patient Name': 'nunique',
            'Claim Code': 'count'
        }).reset_index()
        claims_monthly['Claim_Month'] = claims_monthly['Claim_Month'].astype(str)
        claims_monthly.columns = ['month', 'billed', 'paid', 'patients', 'claims']

        # Parse cost dates and group by month
        costs['Cost_Month'] = pd.to_datetime(costs['Date'], errors='coerce').dt.to_period('M')
        costs_monthly = costs.groupby('Cost_Month')['Total_Amount'].sum().reset_index()
        costs_monthly['Cost_Month'] = costs_monthly['Cost_Month'].astype(str)
        costs_monthly.columns = ['month', 'costs']

        # Merge
        monthly = pd.merge(claims_monthly, costs_monthly, on='month', how='outer').fillna(0)
        monthly = monthly.sort_values('month')
        monthly['profit'] = monthly['paid'] - monthly['costs']

        return {
            "data": monthly.to_dict('records'),
            "summary": {
                "total_billed": float(claims['Claim Amount'].sum()),
                "total_paid": float(claims['Paid Amount'].sum()),
                "total_costs": float(costs['Total_Amount'].sum()),
                "date_range": {
                    "claims_start": str(claims['Claim Start'].min()),
                    "claims_end": str(claims['Claim Start'].max()),
                    "costs_start": str(costs['Date'].min()),
                    "costs_end": str(costs['Date'].max())
                }
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/explorer/physicians")
async def get_physicians_list(current_user: User = Depends(get_current_user)):
    """Get list of physicians for filtering."""
    try:
        claims = pd.read_csv("data/excel/Claim List.csv")
        physicians = claims['Primary Physician'].dropna().unique().tolist()
        return sorted(physicians)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/explorer/export")
async def export_explorer_data(current_user: User = Depends(get_current_user)):
    """Export patient-level summary to Excel."""
    try:
        from profitability_analysis import ProfitabilityAnalyzer

        claims = pd.read_csv("data/excel/Claim List.csv")
        costs = pd.read_excel("data/excel/employee_costs.xlsx")

        # Parse dates
        claims['Claim Start'] = pd.to_datetime(claims['Claim Start'], errors='coerce')
        claims['Final Sent Date'] = pd.to_datetime(claims['Final Sent Date'], errors='coerce')
        costs['Date'] = pd.to_datetime(costs['Date'], errors='coerce')
        costs['Date_Paid'] = pd.to_datetime(costs['Date_Paid'], errors='coerce')

        # ===== SHEET 1: HIGH LEVEL SUMMARY =====
        total_revenue = claims['Paid Amount'].sum()
        total_billed = claims['Claim Amount'].sum()
        total_costs = costs['Total_Amount'].sum()

        summary_data = [
            {'Metric': 'Total Billed', 'Value': total_billed},
            {'Metric': 'Total Revenue (Paid)', 'Value': total_revenue},
            {'Metric': 'Total Costs', 'Value': total_costs},
            {'Metric': 'Net Profit', 'Value': total_revenue - total_costs},
            {'Metric': 'Profit Margin %', 'Value': ((total_revenue - total_costs) / total_revenue * 100) if total_revenue > 0 else 0},
            {'Metric': '', 'Value': ''},
            {'Metric': 'Total Claims', 'Value': len(claims)},
            {'Metric': 'Unique Patients', 'Value': claims['Patient Name'].nunique()},
            {'Metric': 'Unique Physicians', 'Value': claims['Primary Physician'].nunique()},
            {'Metric': '', 'Value': ''},
            {'Metric': 'Claims Date Range', 'Value': f"{claims['Claim Start'].min().strftime('%Y-%m-%d') if pd.notna(claims['Claim Start'].min()) else 'N/A'} to {claims['Claim Start'].max().strftime('%Y-%m-%d') if pd.notna(claims['Claim Start'].max()) else 'N/A'}"},
            {'Metric': 'Costs Date Range', 'Value': f"{costs['Date'].min().strftime('%Y-%m-%d') if pd.notna(costs['Date'].min()) else 'N/A'} to {costs['Date'].max().strftime('%Y-%m-%d') if pd.notna(costs['Date'].max()) else 'N/A'}"},
        ]
        summary_df = pd.DataFrame(summary_data)

        # ===== SHEET 2: PATIENT LEVEL SUMMARY =====
        # Revenue by patient
        patient_revenue = claims.groupby('Patient Name').agg({
            'Claim Amount': 'sum',
            'Paid Amount': 'sum',
            'Balance': 'sum',
            'Claim Code': 'count',
            'Claim Start': 'min',
            'Final Sent Date': 'max',
            'Primary Physician': 'first',
            'Primary Insurance': 'first'
        }).reset_index()
        patient_revenue.columns = ['Patient Name', 'Total Billed', 'Total Paid', 'Balance',
                                   'Claim Count', 'First Claim Date', 'Last Billed Date',
                                   'Primary Physician', 'Insurance']

        # Try to match costs to patients using the profitability analyzer logic
        analyzer = ProfitabilityAnalyzer()
        analyzer.load_data()

        # Build patient cost lookup
        costs['Patient_Name'] = costs['Patient_Name'].astype(str).replace('nan', '')

        # Normalize names for matching
        def normalize(name):
            if pd.isna(name) or name == '' or name == 'nan':
                return ''
            name = str(name).strip().lower()
            name = name.replace('.', '').replace(',', ' ').replace('/', ' ')
            parts = name.split()
            return ' '.join(sorted(parts))

        patient_revenue['norm_name'] = patient_revenue['Patient Name'].apply(normalize)
        costs['norm_name'] = costs['Patient_Name'].apply(normalize)

        # Sum costs by normalized name
        cost_by_norm = costs.groupby('norm_name')['Total_Amount'].sum().to_dict()

        # Match costs
        patient_revenue['Total Costs'] = patient_revenue['norm_name'].map(cost_by_norm).fillna(0)
        patient_revenue['Net Profit'] = patient_revenue['Total Paid'] - patient_revenue['Total Costs']
        patient_revenue['Profit Margin %'] = (patient_revenue['Net Profit'] / patient_revenue['Total Paid'] * 100).round(1)
        patient_revenue.loc[patient_revenue['Total Paid'] == 0, 'Profit Margin %'] = 0

        # Format dates
        patient_revenue['First Claim Date'] = pd.to_datetime(patient_revenue['First Claim Date'], errors='coerce').dt.strftime('%Y-%m-%d')
        patient_revenue['Last Billed Date'] = pd.to_datetime(patient_revenue['Last Billed Date'], errors='coerce').dt.strftime('%Y-%m-%d')

        # Drop norm_name and reorder
        patient_summary = patient_revenue[[
            'Patient Name', 'Primary Physician', 'Insurance',
            'Total Billed', 'Total Paid', 'Total Costs', 'Net Profit', 'Profit Margin %',
            'Balance', 'Claim Count', 'First Claim Date', 'Last Billed Date'
        ]].sort_values('Total Paid', ascending=False)

        # ===== SHEET 3: MONTHLY SUMMARY =====
        claims['Month'] = pd.to_datetime(claims['Claim Start'], errors='coerce').dt.to_period('M').astype(str)
        costs['Month'] = pd.to_datetime(costs['Date'], errors='coerce').dt.to_period('M').astype(str)

        monthly_revenue = claims.groupby('Month').agg({
            'Claim Amount': 'sum',
            'Paid Amount': 'sum',
            'Patient Name': 'nunique',
            'Claim Code': 'count'
        }).reset_index()
        monthly_revenue.columns = ['Month', 'Billed', 'Paid', 'Patients', 'Claims']

        monthly_costs = costs.groupby('Month')['Total_Amount'].sum().reset_index()
        monthly_costs.columns = ['Month', 'Costs']

        monthly_summary = pd.merge(monthly_revenue, monthly_costs, on='Month', how='outer').fillna(0)
        monthly_summary['Net Profit'] = monthly_summary['Paid'] - monthly_summary['Costs']
        monthly_summary = monthly_summary.sort_values('Month')

        # ===== SHEET 4: CLAIMS DETAIL =====
        claims_detail = claims[[
            'Patient Name', 'Primary Physician', 'Primary Insurance',
            'Claim Code', 'Claim Type', 'Status',
            'Claim Start', 'Claim End', 'Final Sent Date',
            'Claim Amount', 'Paid Amount', 'Adjusted Amount', 'Balance'
        ]].copy()
        claims_detail['Claim Start'] = pd.to_datetime(claims_detail['Claim Start'], errors='coerce').dt.strftime('%Y-%m-%d')
        claims_detail['Claim End'] = pd.to_datetime(claims_detail['Claim End'], errors='coerce').dt.strftime('%Y-%m-%d')
        claims_detail['Final Sent Date'] = pd.to_datetime(claims_detail['Final Sent Date'], errors='coerce').dt.strftime('%Y-%m-%d')
        claims_detail = claims_detail.sort_values(['Patient Name', 'Claim Start'])

        # ===== WRITE EXCEL =====
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"outputs/Data_Explorer_Export_{timestamp}.xlsx"
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            patient_summary.to_excel(writer, sheet_name='Patient Summary', index=False)
            monthly_summary.to_excel(writer, sheet_name='Monthly Summary', index=False)
            claims_detail.to_excel(writer, sheet_name='Claims Detail', index=False)

            from openpyxl.styles import numbers

            # Currency columns for each sheet
            currency_formats = {
                'Summary': ['B'],  # Value column for money rows
                'Patient Summary': ['D', 'E', 'F', 'G', 'I'],  # Billed, Paid, Costs, Profit, Balance
                'Monthly Summary': ['B', 'C', 'F', 'G'],  # Billed, Paid, Costs, Net Profit
                'Claims Detail': ['J', 'K', 'L', 'M']  # Claim Amount, Paid, Adjusted, Balance
            }

            # Apply formatting
            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]

                # Auto-fit columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

                # Apply currency format
                if sheet_name in currency_formats:
                    for col_letter in currency_formats[sheet_name]:
                        for row in range(2, ws.max_row + 1):  # Skip header
                            cell = ws[f'{col_letter}{row}']
                            if isinstance(cell.value, (int, float)) and cell.value != 0:
                                cell.number_format = '"$"#,##0.00'

            # Special handling for Summary sheet - only format money rows
            ws = writer.sheets['Summary']
            money_metrics = ['Total Billed', 'Total Revenue (Paid)', 'Total Costs', 'Net Profit']
            for row in range(2, ws.max_row + 1):
                metric_cell = ws[f'A{row}']
                value_cell = ws[f'B{row}']
                if metric_cell.value in money_metrics:
                    if isinstance(value_cell.value, (int, float)):
                        value_cell.number_format = '"$"#,##0.00'
                elif metric_cell.value == 'Profit Margin %':
                    if isinstance(value_cell.value, (int, float)):
                        value_cell.number_format = '0.0"%"'

        return FileResponse(
            path=output_path,
            filename=f"Data_Explorer_Export_{timestamp}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error exporting: {str(e)}")


if __name__ == "__main__":
    print("🚀 Starting Home Health Analytics API Server...")
    print("📊 API Documentation: http://localhost:8000/docs")
    print("🔄 To refresh data: http://localhost:8000/analytics/refresh")

    uvicorn.run(
        "api_server:app",
        host="0.0.0.0",
        port=8000,
        reload=True,
        log_level="info"
    )
